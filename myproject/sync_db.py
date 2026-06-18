"""
Amazon ASIN Product Detail, Excel & Django DB Synchronization Engine
Usage:
    python excel_db_sync.py B0XXXXXXXX B0YYYYYYYY
    python excel_db_sync.py -o custom_output.xlsx B0XXXXXXXX
"""

import os
import sys
import datetime
import time
import re
from decimal import Decimal

# ============================================================
# INITIALIZE DJANGO ENVIRONMENT (Prevents Duplicate Registry)
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(BASE_DIR)
a= os.environ  # <-- Verify this matches your folder name
print(a)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "myproject.settings")  

import django
django.setup()

from django.utils import timezone
from django.db import transaction
from products.models import AmazonLink, Product, ProductSnapshot  # Absolute import from app namespace

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
CREDENTIAL_ID     = "4v6br6lho9jsi7mc1iu418ci5c"
CREDENTIAL_SECRET = "1oga06qok3g9nd2r92ogl5aqpe9b7kumr34e1aolceecdlfpgtnv"
VERSION           = "2.2"
MARKETPLACE       = "www.amazon.in"
PARTNER_TAG       = "alena01-21"

ALL_RESOURCES = [
    "images.primary.small", "images.primary.medium", "images.primary.large", "images.primary.highRes",
    "images.variants.small", "images.variants.medium", "images.variants.large", "images.variants.highRes",
    "itemInfo.byLineInfo", "itemInfo.classifications", "itemInfo.contentInfo", "itemInfo.contentRating",
    "itemInfo.externalIds", "itemInfo.features", "itemInfo.manufactureInfo", "itemInfo.productInfo",
    "itemInfo.technicalInfo", "itemInfo.title", "itemInfo.tradeInInfo",
    "browseNodeInfo.browseNodes", "browseNodeInfo.browseNodes.ancestor", "browseNodeInfo.browseNodes.salesRank",
    "browseNodeInfo.websiteSalesRank", "parentASIN",
    "offersV2.listings.availability", "offersV2.listings.condition", "offersV2.listings.dealDetails",
    "offersV2.listings.isBuyBoxWinner", "offersV2.listings.loyaltyPoints", "offersV2.listings.merchantInfo",
    "offersV2.listings.price", "offersV2.listings.type",
]

# ============================================================
# ACCESS HELPERS
# ============================================================

def ga(obj, *keys, default=None):
    for key in keys:
        if obj is None:
            return default
        snake = re.sub(r'(?<!^)(?=[A-Z])', '_', key).lower()
        val = getattr(obj, snake, None)
        if val is None:
            val = getattr(obj, key, None)
        obj = val
    return obj if obj is not None else default


def dv(obj, default=None):
    if obj is None:
        return default
    v = getattr(obj, "display_value", None)
    if v is None:
        v = getattr(obj, "displayValue", None)
    return v if v is not None else default


def to_number(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None

# ============================================================
# FETCH (API)
# ============================================================

def fetch_product(asin: str):
    from creatorsapi_python_sdk.api_client import ApiClient
    from creatorsapi_python_sdk.api.default_api import DefaultApi
    from creatorsapi_python_sdk.models.get_items_request_content import GetItemsRequestContent
    from creatorsapi_python_sdk.exceptions import ApiException
    try:
        client = ApiClient(
            credential_id=CREDENTIAL_ID,
            credential_secret=CREDENTIAL_SECRET,
            version=VERSION,
        )
        api = DefaultApi(client)
        request = GetItemsRequestContent(
            partner_tag=PARTNER_TAG,
            item_ids=[asin],
            resources=ALL_RESOURCES,
        )
        response = api.get_items(
            x_marketplace=MARKETPLACE,
            get_items_request_content=request,
        )

        if response.errors:
            for err in response.errors:
                print(f"  ⚠  [{err.code}] {err.message}")
            return None
        if not response.items_result or not response.items_result.items:
            print(f"  No item found for ASIN: {asin}")
            return None
        return response.items_result.items[0]
    except ApiException as e:
        print(f"  API Exception for {asin}: {e}")
        return None
    except Exception as e:
        print(f"  Unexpected error for {asin}: {e}")
        return None

# ============================================================
# EXTRACT (Item Object -> Plain Dicts Map)
# ============================================================

def _dim_str(d):
    if not d: return None
    val = getattr(d, "display_value", None)
    unit = getattr(d, "unit", "") or ""
    if val is None: return None
    return f"{val} {unit}".strip()


def _join_display_values(container):
    if not container or not hasattr(container, "display_values") or not container.display_values:
        return None
    out = []
    for v in container.display_values:
        out.append(dv(v) if hasattr(v, "display_value") else str(v))
    return ", ".join(str(x) for x in out if x)


def extract_product(item):
    asin = getattr(item, "asin", None)
    url = (getattr(item, "detail_page_url", None) or f"https://www.amazon.in/dp/{asin}?tag={PARTNER_TAG}")

    flat = {"asin": asin, "url": url}
    info = getattr(item, "item_info", None)
    flat["title"] = dv(ga(info, "title"))

    # byLineInfo
    bli = ga(info, "byLineInfo")
    flat["brand"] = dv(ga(bli, "brand"))
    flat["manufacturer"] = dv(ga(bli, "manufacturer"))
    contribs = getattr(bli, "contributors", None) if bli else None
    contrib_list = []
    if contribs:
        for c in contribs:
            name = getattr(c, "name", "") or ""
            role = getattr(c, "role", "") or getattr(c, "role_type", "") or getattr(c, "roleType", "")
            if name: contrib_list.append({"name": name, "role": role})
    flat["contributors_dict"] = contrib_list
    flat["contributors"] = ", ".join([f"{c['name']} ({c['role']})" if c['role'] else c['name'] for c in contrib_list]) if contrib_list else None

    # Classifications
    cls = ga(info, "classifications")
    flat["product_group"] = dv(ga(cls, "productGroup"))
    flat["binding"] = dv(ga(cls, "binding"))

    # Manufacture Info
    mi = ga(info, "manufactureInfo")
    flat["part_number"] = dv(ga(mi, "itemPartNumber"))
    flat["model"] = dv(ga(mi, "model"))
    flat["warranty"] = dv(ga(mi, "warranty"))

    # Product Info & Dimensions
    pi = ga(info, "productInfo")
    flat["color"] = dv(ga(pi, "color"))
    flat["size"] = dv(ga(pi, "size"))
    flat["unit_count"] = dv(ga(pi, "unitCount"))
    flat["release_date"] = dv(ga(pi, "releaseDate"))
    
    dims = ga(pi, "itemDimensions")
    flat["dim_h"] = dv(ga(dims, "height"))
    flat["dim_l"] = dv(ga(dims, "length"))
    flat["dim_w"] = dv(ga(dims, "width"))
    flat["dim_wt"] = dv(ga(dims, "weight"))
    if dims:
        hlw = [s for s in [_dim_str(ga(dims, "height")), _dim_str(ga(dims, "length")), _dim_str(ga(dims, "width"))] if s]
        flat["dimensions"] = " × ".join(hlw) if hlw else None
        flat["weight"] = _dim_str(ga(dims, "weight"))
    else:
        flat["dimensions"] = None
        flat["weight"] = None

    # External IDs
    ext = ga(info, "externalIds")
    flat["ean"] = _join_display_values(getattr(ext, "eans", None)) if ext else None
    flat["upc"] = _join_display_values(getattr(ext, "upcs", None)) if ext else None
    flat["isbn"] = _join_display_values(getattr(ext, "isbns", None)) if ext else None

    # Features
    feats_obj = ga(info, "features")
    features = []
    if feats_obj and hasattr(feats_obj, "display_values") and feats_obj.display_values:
        for f in feats_obj.display_values:
            features.append(dv(f) if hasattr(f, "display_value") else str(f))
    flat["features_list"] = features
    flat["features_count"] = len(features)

    # Images
    images_obj = getattr(item, "images", None)
    flat["image_url"] = None
    flat["image_dims"] = None
    flat["variant_count"] = 0
    flat["variant_images"] = []
    if images_obj:
        primary = getattr(images_obj, "primary", None)
        if primary:
            for size in ["large", "high_res", "highRes", "medium", "small"]:
                img = getattr(primary, size, None)
                u = getattr(img, "url", None) if img else None
                if u:
                    flat["image_url"] = u
                    flat["image_w"] = getattr(img, "width", None)
                    flat["image_h"] = getattr(img, "height", None)
                    flat["image_dims"] = f"{flat['image_w']}x{flat['image_h']}" if flat['image_w'] and flat['image_h'] else None
                    break
        variants = getattr(images_obj, "variants", None)
        if variants:
            flat["variant_count"] = len(variants)
            for v in variants:
                for size in ["large", "high_res", "highRes", "medium", "small"]:
                    v_meta = getattr(v, size, None)
                    if v_meta and getattr(v_meta, "url", None):
                        flat["variant_images"].append(v_meta.url)
                        break

    flat["parent_asin"] = getattr(item, "parent_asin", None)

    # Rankings
    rankings = []
    flat["overall_rank"] = None
    flat["overall_rank_context"] = None
    bni = getattr(item, "browse_node_info", None)
    if bni:
        wsr = getattr(bni, "website_sales_rank", None)
        if wsr:
            flat["overall_rank"] = getattr(wsr, "sales_rank", None)
            flat["overall_rank_context"] = (getattr(wsr, "context_free_name", None) or getattr(wsr, "display_name", None))
        nodes = getattr(bni, "browse_nodes", None) or []
        for n in nodes:
            rank = getattr(n, "sales_rank", None)
            if not rank: continue
            name = (getattr(n, "context_free_name", None) or getattr(n, "display_name", "Unknown"))
            chain = []
            anc = getattr(n, "ancestor", None)
            while anc:
                aname = (getattr(anc, "context_free_name", None) or getattr(anc, "display_name", ""))
                if aname: chain.append(aname)
                anc = getattr(anc, "ancestor", None)
            path = " > ".join(reversed(chain) + [name]) if chain else name
            rankings.append({
                "node_id": str(getattr(n, "id", "")), "name": name, "rank": int(rank),
                "is_root": getattr(n, "is_root", False), "path": path,
            })

    # Offers (All Listings)
    offers_rows = []
    offers = getattr(item, "offers_v2", None)
    listings = getattr(offers, "listings", None) if offers else None
    if listings:
        for idx, lst in enumerate(listings, 1):
            o = {"asin": asin, "listing": idx}
            price_obj = getattr(lst, "price", None)
            money = getattr(price_obj, "money", None) if price_obj else None
            o["price_display"] = getattr(money, "display_amount", None) if money else None
            o["price_amount"] = to_number(getattr(money, "amount", None)) if money else None
            o["currency"] = getattr(money, "currency", None) if money else None

            savings = getattr(price_obj, "savings", None) if price_obj else None
            o["discount_pct"] = getattr(savings, "percentage", None) if savings else None
            sav_money = getattr(savings, "money", None) if savings else None
            o["savings_display"] = getattr(sav_money, "display_amount", None) if sav_money else None

            sb = getattr(price_obj, "saving_basis", None) if price_obj else None
            sb_money = getattr(sb, "money", None) if sb else None
            o["mrp_label"] = (getattr(sb, "saving_basis_type_label", None) or "MRP") if sb else None
            o["mrp_display"] = getattr(sb_money, "display_amount", None) if sb_money else None
            o["mrp_amount"] = to_number(getattr(sb_money, "amount", None)) if sb_money else None

            cond = getattr(lst, "condition", None)
            o["condition"] = (getattr(cond, "value", None) or getattr(cond, "display_value", None)) if cond else None

            avail = getattr(lst, "availability", None)
            o["availability"] = getattr(avail, "message", None) if avail else None
            o["availability_type"] = getattr(avail, "type", None) if avail else None

            merchant = getattr(lst, "merchant_info", None)
            o["merchant_name"] = getattr(merchant, "name", None) if merchant else None
            o["merchant_id"] = getattr(merchant, "id", None) if merchant else None
            o["merchant"] = (o["merchant_name"] or o["merchant_id"])

            o["buy_box"] = getattr(lst, "is_buy_box_winner", None)
            o["listing_type"] = getattr(lst, "type", None)

            deal = getattr(lst, "deal_details", None)
            o["deal_type"] = getattr(deal, "type", None) if deal else None
            o["deal_end"] = getattr(deal, "end_time", None)

            lp = getattr(lst, "loyalty_points", None)
            o["loyalty_points"] = getattr(lp, "points", None) if lp else None

            offers_rows.append(o)

    # Hydrate primary values onto flat dict maps
    primary_offer = next((o for o in offers_rows if o.get("buy_box")), None) or (offers_rows[0] if offers_rows else {})
    for k in ["price_display", "price_amount", "currency", "discount_pct", "savings_display", "savings_amount",
              "mrp_label", "mrp_display", "mrp_amount", "condition", "availability", "availability_type",
              "merchant", "merchant_name", "merchant_id", "buy_box", "listing_type", "deal_type", "deal_end", "loyalty_points"]:
        flat[k] = primary_offer.get(k)

    return {
        "flat": flat,
        "rankings": rankings,
        "features": [{"asin": asin, "title": flat["title"], "n": i + 1, "feature": f} for i, f in enumerate(features)],
        "offers": offers_rows,
    }

# ============================================================
# ACCURATE DJANGO ORM PERSISTENCE WRAPPER
# ============================================================

def save_to_django_db(data):
    """Accurately maps the structured API fields directly into the database models."""
    flat = data["flat"]
    asin = flat["asin"]
    
    with transaction.atomic():
        # 1. Update/Create Link Layer Anchor Row
        link, created = AmazonLink.objects.get_or_create(asin=asin)
        link.product_url = flat["url"]
        link.title = (flat["title"] or "Amazon Product")[:200]
        link.save()

        # 2. Update/Create Product Dashboard Data Metric Snap
        product, _ = Product.objects.get_or_create(link=link)
        product.title = (flat["title"] or "")[:500]
        product.description = f"Latest automated catalog synchronization asset check loop for ASIN: {asin}."
        
        # Dimensions & Meta properties mapping
        product.image_url = (flat.get("image_url") or "")[:2000]
        product.image_width = flat.get("image_w")
        product.image_height = flat.get("image_h")
        product.variant_images = flat.get("variant_images") or []
        
        product.brand = (flat.get("brand") or "")[:255]
        product.manufacturer = (flat.get("manufacturer") or "")[:255]
        product.contributors = flat.get("contributors_dict") or []
        product.product_group = (flat.get("product_group") or "")[:255]
        product.binding = (flat.get("binding") or "")[:255]
        
        product.item_part_number = (flat.get("part_number") or "")[:255]
        product.model_number = (flat.get("model") or "")[:255]
        product.warranty = flat.get("warranty") or ""
        
        product.color = (flat.get("color") or "")[:255]
        product.size = (flat.get("size") or "")[:255]
        product.unit_count = (flat.get("unit_count") or "")[:100]
        
        product.dimension_height = (flat.get("dim_h") or "")[:50]
        product.dimension_length = (flat.get("dim_l") or "")[:50]
        product.dimension_width = (flat.get("dim_w") or "")[:50]
        product.dimension_weight = (flat.get("dim_wt") or "")[:50]
        
        # Lists and structural nodes
        product.features = flat.get("features_list") or []
        product.parent_asin = (flat.get("parent_asin") or "")[:10]
        product.overall_rank = flat.get("overall_rank")
        product.overall_rank_context = (flat.get("overall_rank_context") or "")[:255]
        product.category_rankings = data["rankings"]

        # Offer parameters mapping
        if flat.get("price_amount") is not None:
            product.price_amount = Decimal(str(flat["price_amount"]))
        product.price_display = (flat.get("price_display") or "")[:50]
        product.price_currency = (flat.get("currency") or "INR")[:10]

        if flat.get("mrp_amount") is not None:
            product.mrp_amount = Decimal(str(flat["mrp_amount"]))
        product.mrp_display = (flat.get("mrp_display") or "")[:50]
        product.mrp_label = (flat.get("mrp_label") or "MRP")[:100]

        product.discount_percentage = flat.get("discount_pct")
        if flat.get("savings_amount") is not None:
            product.savings_amount = Decimal(str(flat["savings_amount"]))
        product.savings_display = (flat.get("savings_display") or "")[:50]

        product.condition = (flat.get("condition") or "")[:100]
        product.availability_message = (flat.get("availability") or "")[:255]
        product.availability_type = (flat.get("availability_type") or "")[:100]
        product.merchant_name = (flat.get("merchant_name") or "")[:255]
        product.merchant_id = (flat.get("merchant_id") or "")[:255]
        product.is_buy_box_winner = flat.get("buy_box")
        product.listing_type = (flat.get("listing_type") or "")[:100]
        product.deal_type = (flat.get("deal_type") or "")[:100]
        product.deal_end_time = flat.get("deal_end")
        product.loyalty_points = flat.get("loyalty_points")
        
        product.last_checked_at = timezone.now()
        product.save()

        # 3. Append Chronic Performance Tracking Historical Snapshot Row
        ProductSnapshot.objects.create(
            product=product,
            price_amount=product.price_amount,
            price_display=product.price_display,
            mrp_amount=product.mrp_amount,
            mrp_display=product.mrp_display,
            discount_percentage=product.discount_percentage,
            savings_amount=product.savings_amount,
            availability_message=product.availability_message,
            availability_type=product.availability_type,
            merchant_name=product.merchant_name,
            is_buy_box_winner=product.is_buy_box_winner,
            overall_rank=product.overall_rank,
            deal_type=product.deal_type,
            deal_end_time=product.deal_end_time
        )
    print(f"   💾 Database Sync Active: Saved model structures securely into DB space.")

# ============================================================
# EXCEL WRITER
# ============================================================

HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="Calibri", size=11)
LINK_FONT   = Font(name="Calibri", size=11, color="2563EB", underline="single")
THIN        = Side(style="thin", color="E5E7EB")
BORDER      = Border(bottom=THIN)

PRODUCTS_COLS = [
    ("ASIN", "asin", 14, "text"), ("Title", "title", 50, "text"),
    ("Brand", "brand", 18, "text"), ("Manufacturer", "manufacturer", 18, "text"),
    ("Contributors", "contributors", 24, "text"), ("Product Group", "product_group", 18, "text"),
    ("Binding", "binding", 14, "text"), ("Part Number", "part_number", 18, "text"),
    ("Model", "model", 16, "text"), ("Warranty", "warranty", 20, "text"),
    ("Color", "color", 14, "text"), ("Size", "size", 14, "text"),
    ("Unit Count", "unit_count", 11, "text"), ("Dimensions (H×L×W)", "dimensions", 22, "text"),
    ("Weight", "weight", 14, "text"), ("Price (display)", "price_display", 14, "text"),
    ("Price (amount)", "price_amount", 14, "money"), ("Currency", "currency", 10, "text"),
    ("MRP label", "mrp_label", 11, "text"), ("MRP (display)", "mrp_display", 14, "text"),
    ("Discount %", "discount_pct", 11, "pct"), ("Savings", "savings_display", 14, "text"),
    ("Condition", "condition", 12, "text"), ("Availability", "availability", 24, "text"),
    ("Availability Type", "availability_type", 16, "text"), ("Merchant", "merchant", 20, "text"),
    ("Buy Box Winner", "buy_box", 14, "text"), ("Listing Type", "listing_type", 14, "text"),
    ("Deal Type", "deal_type", 14, "text"), ("Deal End", "deal_end", 20, "text"),
    ("Loyalty Points", "loyalty_points", 14, "num"), ("Overall Site Rank", "overall_rank", 16, "num"),
    ("Overall Rank Context", "overall_rank_context", 24, "text"), ("EAN", "ean", 16, "text"),
    ("UPC", "upc", 16, "text"), ("ISBN", "isbn", 16, "text"),
    ("Parent ASIN", "parent_asin", 14, "text"), ("Features Count", "features_count", 13, "num"),
    ("Image WxH", "image_dims", 12, "text"), ("Variants", "variant_count", 10, "num"),
    ("Primary Image", "image_url", 40, "url"), ("Detail URL", "url", 40, "url"),
]

OFFERS_COLS = [
    ("ASIN", "asin", 14, "text"), ("Listing #", "listing", 10, "num"),
    ("Price (display)", "price_display", 14, "text"), ("Price (amount)", "price_amount", 14, "money"),
    ("Currency", "currency", 10, "text"), ("Discount %", "discount_pct", 11, "pct"),
    ("Savings", "savings_display", 14, "text"), ("MRP label", "mrp_label", 11, "text"),
    ("MRP (display)", "mrp_display", 14, "text"), ("Condition", "condition", 12, "text"),
    ("Availability", "availability", 24, "text"), ("Availability Type", "availability_type", 16, "text"),
    ("Merchant", "merchant", 20, "text"), ("Buy Box Winner", "buy_box", 14, "text"),
    ("Listing Type", "listing_type", 14, "text"), ("Deal Type", "deal_type", 14, "text"),
    ("Deal End", "deal_end", 20, "text"), ("Loyalty Points", "loyalty_points", 14, "num"),
]

RANKINGS_COLS = [
    ("ASIN", "asin", 14, "text"), ("Title", "title", 45, "text"), ("Category", "category", 28, "text"),
    ("Rank", "rank", 12, "num"), ("Is Root", "is_root", 10, "text"), ("Node ID", "node_id", 16, "text"),
    ("Path", "path", 60, "text"),
]

FEATURES_COLS = [
    ("ASIN", "asin", 14, "text"), ("Title", "title", 45, "text"), ("#", "n", 6, "num"), ("Feature", "feature", 90, "text"),
]

def _write_sheet(ws, columns, rows):
    for c, (header, _key, width, _kind) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 20

    for r, rowdata in enumerate(rows, 2):
        for c, (_header, key, _width, kind) in enumerate(columns, 1):
            val = rowdata.get(key)
            cell = ws.cell(row=r, column=c)
            if kind == "url" and val:
                cell.value = val
                cell.hyperlink = val
                cell.font = LINK_FONT
            else:
                if kind in ("money", "num", "pct"):
                    num = to_number(val)
                    cell.value = num if num is not None else (val if val not in (None, "") else None)
                elif isinstance(val, bool):
                    cell.value = "Yes" if val else "No"
                else:
                    cell.value = val
                cell.font = BODY_FONT
                if kind == "money" and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                elif kind == "pct" and isinstance(cell.value, (int, float)):
                    cell.number_format = '0"%"'
                elif kind == "num" and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
            cell.alignment = Alignment(vertical="top", wrap_text=(kind == "text" and _width >= 40))
            cell.border = BORDER

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"


def write_excel(extracted, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    _write_sheet(ws, PRODUCTS_COLS, [e["flat"] for e in extracted])
    _write_sheet(wb.create_sheet("Offers"), OFFERS_COLS, [o for e in extracted for o in e["offers"]])
    _write_sheet(wb.create_sheet("Category Rankings"), RANKINGS_COLS, [r for e in extracted for r in e["rankings"]])
    _write_sheet(wb.create_sheet("Features"), FEATURES_COLS, [f for e in extracted for f in e["features"]])
    wb.save(path)


# ============================================================
# ENTRY POINT
# ============================================================

def main():
    args = sys.argv[1:]
    out_path = None
    if "-o" in args:
        i = args.index("-o")
        try:
            out_path = args[i + 1]
            del args[i:i + 2]
        except IndexError:
            print("-o requires a filename")
            return

    if args:
        asins = args
    else:
        raw = input("\nEnter ASIN(s) separated by spaces: ").strip()
        asins = raw.split()

    if not asins:
        print("No ASINs provided. Exiting.")
        return

    if not out_path:
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"amazon_products_{stamp}.xlsx"

    print(f"\nFetching {len(asins)} ASIN(s) from {MARKETPLACE} ...\n")

    extracted = []
    for asin in asins:
        asin = asin.strip().upper()
        if len(asin) != 10:
            print(f"  Skipping '{asin}' — ASINs must be exactly 10 characters")
            continue
        
        item = fetch_product(asin)
        if item is None:
            continue
            
        data = extract_product(item)
        extracted.append(data)
        
        # Save to real production database row matching standard limits
        save_to_django_db(data)
        
        f = data["flat"]
        print(f"  ✓ {f['asin']}  {(f.get('title') or '')[:55]:<55}  {f.get('price_display') or '-'}")
        
        # ALWAYS PAUSE 2 SECONDS AFTER BOTH DB WRITING AND CONCURRENT RETRIEVALS
        print("   ⏳ Throttling break: Pausing execution path for 2 seconds...")
        time.sleep(2)

    if not extracted:
        print("\nNothing fetched — no Excel file written.")
        return

    write_excel(extracted, out_path)
    print(f"\n✅ Saved {len(extracted)} product(s) to: {out_path}")
    print("   Sheets: Products | Offers | Category Rankings | Features")


if __name__ == "__main__":
    main()